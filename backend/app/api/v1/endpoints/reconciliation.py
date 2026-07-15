"""
app/api/v1/endpoints/remaining_inventory.py
Remaining Inventory dashboard endpoints

Opening & Closing Stock Methodology
─────────────────────────────────────
All stock values are reconstructed from InventoryTransaction rows.
quantity is always stored as a positive number; direction is determined
by transaction_type:

    PURCHASE                → stock IN  (+qty)
    PREPARATION / SALE
    WASTAGE / ADJUSTMENT    → stock OUT (−qty)

Opening stock at period start  = SUM of all signed transaction values
                                  with transaction_date < start_dt

Closing stock at period end    = Opening + SUM of signed transaction values
                                  with transaction_date BETWEEN start_dt AND end_dt

This gives accurate point-in-time values regardless of what happened to
quantity_remaining after the fact.

COGS Methodology
─────────────────
COGS uses the standard restaurant formula:

    COGS = Opening Stock + Purchases − Closing Stock

This captures ALL inventory consumption automatically — sales, preparation,
wastage, spillage, theft, and any unrecorded usage. Wastage is also tracked
separately as an informational breakdown inside COGS.

Food Cost % = (COGS / Revenue) × 100
Healthy restaurant range: 28–35%.

NOTE ON WASTAGE:
Wastage quantities and values are read from inventory_transactions
(TransactionType.WASTAGE) and reported as an informational subset of COGS.

UNIT CONVERSION:
─────────────────
Transactions may be recorded in a different unit than the inventory item's
base unit (e.g. a purchase entered in "gm" while the item's base unit is "kg").
All quantities are normalized to the item's base unit before aggregation using
the UNIT_CONVERSION_FACTORS table. If no conversion path is found, the quantity
is used as-is and a warning is logged.

Supported conversions (bidirectional):
  Weight : kg ↔ gm ↔ mg
  Volume : liter ↔ ml
  Others : pcs, packet, bottle, rupee, m, sheet — treated as identical units
"""

from math import ceil

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import func, case
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date, datetime, timedelta, timezone
from enum import Enum
import logging

from app.api.deps import get_db
from app.models.inventory import Inventory, InventoryBatch, InventoryTransaction, TransactionType
from app.models.users import User
from app.utils.auth_helper import get_current_user
import io
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo


router = APIRouter()
logger = logging.getLogger(__name__)

STOCK_OUT_TYPES = (
    TransactionType.PREPARATION,
    TransactionType.SALE,
    TransactionType.WASTAGE,
    TransactionType.ADJUSTMENT,
)

MAX_CUSTOM_RANGE_DAYS = 90


# ──────────────────────────────────────────────────────────────────────────────
# Unit Conversion
# ──────────────────────────────────────────────────────────────────────────────

# All values relative to a common base:
#   Weight  → gm  (1 kg = 1000 gm, 1 mg = 0.001 gm)
#   Volume  → ml  (1 liter = 1000 ml)
#   Others  → themselves (no conversion needed)
UNIT_TO_BASE: dict[str, tuple[str, float]] = {
    # Weight
    "mg":    ("gm", 0.001),
    "gm":    ("gm", 1.0),
    "kg":    ("gm", 1000.0),

    # Volume
    "ml":    ("ml", 1.0),
    "liter": ("ml", 1000.0),
    "l":     ("ml", 1000.0),

    # Discrete / currency — no meaningful conversion, treat as same family
    "pcs":    ("pcs",    1.0),
    "packet": ("packet", 1.0),
    "bottle": ("bottle", 1.0),
    "rupee":  ("rupee",  1.0),
    "m":      ("m",      1.0),
    "sheet":  ("sheet",  1.0),
}


def _conversion_factor(from_unit: str, to_unit: str) -> float:
    """
    Return the multiplication factor to convert a quantity
    from `from_unit` to `to_unit`.

    Examples:
        _conversion_factor("gm", "kg")  → 0.001
        _conversion_factor("kg", "gm")  → 1000.0
        _conversion_factor("kg", "kg")  → 1.0
        _conversion_factor("gm", "ml")  → 1.0  (different families → no-op)

    If the units belong to different families (e.g. gm vs ml) or are
    unrecognised, 1.0 is returned and a warning is logged so existing
    behaviour is preserved rather than silently corrupting data.
    """
    from_unit = (from_unit or "").lower().strip()
    to_unit   = (to_unit   or "").lower().strip()

    if from_unit == to_unit:
        return 1.0

    from_info = UNIT_TO_BASE.get(from_unit)
    to_info   = UNIT_TO_BASE.get(to_unit)

    if from_info is None or to_info is None:
        logger.warning(
            "Unit conversion: unrecognised unit(s) '%s' → '%s'. Using factor 1.0.",
            from_unit, to_unit,
        )
        return 1.0

    from_base, from_factor = from_info
    to_base,   to_factor   = to_info

    if from_base != to_base:
        logger.warning(
            "Unit conversion: incompatible families '%s' (%s) → '%s' (%s). Using factor 1.0.",
            from_unit, from_base, to_unit, to_base,
        )
        return 1.0

    # factor = from_factor / to_factor
    # e.g. gm→kg: 1.0 / 1000.0 = 0.001
    return from_factor / to_factor


def _normalize_qty(quantity: float, from_unit: str, to_unit: str) -> float:
    """Convert `quantity` from `from_unit` to `to_unit`."""
    return quantity * _conversion_factor(from_unit, to_unit)


def _normalize_value(value: float, from_unit: str, to_unit: str) -> float:
    """
    Re-scale a monetary value when the transaction unit differs from the
    item's base unit.

    If 500 gm was purchased for ₹30 and the base unit is kg, the value per kg
    is ₹60 — but the *total monetary value* of those 500 gm is still ₹30.
    We do NOT rescale the monetary value; only the quantity needs conversion.
    """
    return value  # monetary value is always correct as stored


# ──────────────────────────────────────────────────────────────────────────────
# Period helpers
# ──────────────────────────────────────────────────────────────────────────────

class PeriodType(str, Enum):
    DAILY   = "daily"
    WEEKLY  = "weekly"
    MONTHLY = "monthly"
    CUSTOM  = "custom"


def get_period_range(
    period: PeriodType,
    reference_date: Optional[date] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> tuple[datetime, datetime]:
    """
    Return (start_dt, end_dt) for the requested period.
    Both datetimes are timezone-aware (UTC).
    """
    today     = reference_date or datetime.now(timezone.utc).date()
    today_utc = datetime.now(timezone.utc).date()

    if period == PeriodType.DAILY:
        start = datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc)
        end   = datetime.combine(today, datetime.max.time()).replace(tzinfo=timezone.utc)

    elif period == PeriodType.WEEKLY:
        week_start = today - timedelta(days=today.weekday())
        week_end   = week_start + timedelta(days=6)
        start = datetime.combine(week_start, datetime.min.time()).replace(tzinfo=timezone.utc)
        end   = datetime.combine(week_end,   datetime.max.time()).replace(tzinfo=timezone.utc)

    elif period == PeriodType.MONTHLY:
        start = datetime.combine(today.replace(day=1), datetime.min.time()).replace(tzinfo=timezone.utc)
        if today.month == 12:
            last_day = today.replace(day=31)
        else:
            last_day = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        end = datetime.combine(last_day, datetime.max.time()).replace(tzinfo=timezone.utc)

    elif period == PeriodType.CUSTOM:
        if not start_date or not end_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Both start_date and end_date are required when period=custom.",
            )
        if start_date > end_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="start_date must not be after end_date.",
            )
        if start_date > today_utc or end_date > today_utc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="start_date and end_date cannot be in the future.",
            )
        if (end_date - start_date).days > MAX_CUSTOM_RANGE_DAYS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Custom date range cannot exceed {MAX_CUSTOM_RANGE_DAYS} days. "
                    f"Requested: {(end_date - start_date).days} days."
                ),
            )
        start = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end   = datetime.combine(end_date,   datetime.max.time()).replace(tzinfo=timezone.utc)

    return start, end


def _resolve_unit(unit) -> str:
    """Safely extract string value from unit — handles plain str or Enum."""
    return unit.value if hasattr(unit, "value") else str(unit)


# ──────────────────────────────────────────────────────────────────────────────
# Shared helper: resolve total value for a single transaction
# ──────────────────────────────────────────────────────────────────────────────

def _tx_value(tx) -> float:
    """
    Return the correct monetary value for a transaction.

    Priority:
      1. total_value  — set by the FIFO batch calculation; always correct,
                        even for fixed_cost items where unit_cost stores
                        the fixed_cost_amount rather than a per-unit price.
      2. quantity × unit_cost — fallback for legacy rows that predate the
                        total_value column being populated.
    """
    total_value = float(tx.total_value or 0)
    qty_x_cost  = float((tx.quantity or 0) * (tx.unit_cost or 0))
    return total_value if total_value != 0 else qty_x_cost


# ──────────────────────────────────────────────────────────────────────────────
# Python-side aggregation helpers (replaces raw SQL SUM for unit-aware totals)
# ──────────────────────────────────────────────────────────────────────────────

def _aggregate_transactions(
    transactions: list,
    base_unit: str,
    tx_types_positive: tuple,
) -> dict[int, dict]:
    """
    Aggregate a list of InventoryTransaction rows into per-item
    {qty, value} dicts, normalising each transaction's quantity to
    the item's base_unit before summing.

    `tx_types_positive` — transaction types that increase stock (PURCHASE).
    All other types decrease stock.

    Returns:
        { item_id: {"qty": float, "value": float} }

    NOTE: base_unit is passed per-call from the Inventory.unit field.
    For the bulk queries we pass a lookup dict instead — see callers.
    """
    raise NotImplementedError("Use _aggregate_transactions_bulk instead.")


def _aggregate_transactions_bulk(
    rows: list,
    item_base_unit_map: dict[int, str],
    positive_types: tuple,
) -> dict[int, dict]:
    result: dict[int, dict] = {}
    for tx in rows:
        item_id   = tx.inventory_item_id
        base_unit = item_base_unit_map.get(item_id, "")
        tx_unit   = _resolve_unit(tx.unit) if hasattr(tx, "unit") and tx.unit else base_unit

        norm_qty = _normalize_qty(float(tx.quantity or 0), tx_unit, base_unit)

        #  Use shared _tx_value helper — consistent across all aggregations
        value = _tx_value(tx)

        sign = 1 if tx.transaction_type in positive_types else -1

        if item_id not in result:
            result[item_id] = {"qty": 0.0, "value": 0.0}
        result[item_id]["qty"]   += sign * norm_qty
        result[item_id]["value"] += sign * value

    return result


# ──────────────────────────────────────────────────────────────────────────────
# /summary
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/summary", status_code=status.HTTP_200_OK)
def get_remaining_inventory_summary(
    period: PeriodType = Query(PeriodType.MONTHLY),
    reference_date: Optional[date] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.tenant_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tenant access required")

    if period != PeriodType.CUSTOM and (start_date or end_date):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="start_date and end_date are only valid when period=custom.",
        )

    start_dt, end_dt = get_period_range(period, reference_date, start_date, end_date)
    tenant_id = current_user.tenant_id

    try:
        # ── Fetch all active inventory items for this tenant ──────────────────
        items = (
            db.query(Inventory)
            .filter(Inventory.tenant_id == tenant_id, Inventory.is_active == True)
            .all()
        )
        item_ids = [i.id for i in items]
        item_base_unit_map = {i.id: _resolve_unit(i.unit) for i in items}

        if not item_ids:
            return _empty_summary_response(period, start_dt, end_dt)

        # ── Fetch raw transactions BEFORE period (for opening stock) ──────────
        opening_txs = (
            db.query(InventoryTransaction)
            .filter(
                InventoryTransaction.tenant_id == tenant_id,
                InventoryTransaction.inventory_item_id.in_(item_ids),
                InventoryTransaction.transaction_date < start_dt,
            )
            .all()
        )

        # ── Fetch raw transactions WITHIN period ──────────────────────────────
        period_txs = (
            db.query(InventoryTransaction)
            .filter(
                InventoryTransaction.tenant_id == tenant_id,
                InventoryTransaction.inventory_item_id.in_(item_ids),
                InventoryTransaction.transaction_date >= start_dt,
                InventoryTransaction.transaction_date <= end_dt,
            )
            .all()
        )

        # ── Aggregate with unit normalisation ─────────────────────────────────
        opening_map = _aggregate_transactions_bulk(
            opening_txs, item_base_unit_map, positive_types=(TransactionType.PURCHASE,)
        )
        period_map = _aggregate_transactions_bulk(
            period_txs, item_base_unit_map, positive_types=(TransactionType.PURCHASE,)
        )

        # ── Opening / closing stock totals ────────────────────────────────────
        opening_stock = max(sum(v["value"] for v in opening_map.values()), 0.0)
        period_value  = sum(v["value"] for v in period_map.values())
        closing_stock = max(opening_stock + period_value, 0.0)

        # ── Purchases total in period ─────────────────────────────────────────
        #  FIX: use _tx_value() so each purchase transaction uses total_value
        # (FIFO-based), falling back to qty*cost only for legacy rows.
        # This ensures that adding more stock to an item correctly accumulates
        # the running total — e.g. rice ₹1000 + ₹300 = ₹1300 highest expense.
        purchase_txs = [tx for tx in period_txs if tx.transaction_type == TransactionType.PURCHASE]
        purchases_by_item: dict[int, float] = {}
        for tx in purchase_txs:
            item_id = tx.inventory_item_id
            value   = _tx_value(tx)  #  was: float((tx.quantity or 0) * (tx.unit_cost or 0))
            purchases_by_item[item_id] = purchases_by_item.get(item_id, 0.0) + value
        purchases_in_period = sum(purchases_by_item.values())

        # ── Highest / Lowest expense item (by purchase cost in period) ────────
        item_name_map = {i.id: i.name for i in items}

        highest_expense_item  = None
        highest_expense_value = None
        lowest_expense_item   = None
        lowest_expense_value  = None

        active_purchases = {
            item_id: value
            for item_id, value in purchases_by_item.items()
            if value > 0
        }

        if len(active_purchases) >= 1:
            highest_id            = max(active_purchases, key=active_purchases.get)
            highest_expense_item  = item_name_map.get(highest_id)
            highest_expense_value = round(active_purchases[highest_id], 2)

        if len(active_purchases) >= 2:
            lowest_id            = min(active_purchases, key=active_purchases.get)
            lowest_expense_item  = item_name_map.get(lowest_id)
            lowest_expense_value = round(active_purchases[lowest_id], 2)

        # ── Wastage in period ─────────────────────────────────────────────────
        #  FIX: use _tx_value() for consistency with fixed_cost items.
        # Previously used qty*unit_cost which is incorrect for fixed_cost items.
        wastage_txs = [tx for tx in period_txs if tx.transaction_type == TransactionType.WASTAGE]
        wastage = sum(_tx_value(tx) for tx in wastage_txs)  #  was: qty * unit_cost

        # ── Revenue (SALE transactions) ───────────────────────────────────────
        sale_txs = [tx for tx in period_txs if tx.transaction_type == TransactionType.SALE]
        revenue  = sum(float(tx.total_value or 0) for tx in sale_txs)

        # ── Most frequently purchased item ────────────────────────────────────
        purchase_count: dict[int, int] = {}
        for tx in purchase_txs:
            purchase_count[tx.inventory_item_id] = purchase_count.get(tx.inventory_item_id, 0) + 1

        most_frequent_purchased = None
        most_frequent_value     = None

        if purchase_count:
            freq_id                 = max(purchase_count, key=purchase_count.get)
            most_frequent_purchased = item_name_map.get(freq_id)
            most_frequent_value     = round(purchases_by_item.get(freq_id, 0.0), 2)

        # ── COGS & Food Cost % ────────────────────────────────────────────────
        cogs = round(opening_stock + purchases_in_period - closing_stock, 2)
        food_cost_percentage = round((cogs / revenue) * 100, 2) if revenue > 0 else 0.0

        has_activity = any([purchases_in_period > 0, wastage > 0, revenue > 0, cogs > 0])

        return {
            "success": True,
            "message": "Remaining inventory summary fetched successfully",
            "period": period.value,
            "period_range": {
                "start": start_dt.isoformat(),
                "end":   end_dt.isoformat(),
            },
            "has_activity": has_activity,
            "data": {
                "opening_stock":                     opening_stock,
                "closing_stock":                     closing_stock,
                "purchases":                         purchases_in_period,
                "cogs":                              cogs,
                "wastage":                           wastage,
                "revenue":                           revenue,
                "food_cost_percentage":              food_cost_percentage,
                "highest_expense_item":              highest_expense_item,
                "highest_expense_value":             highest_expense_value,
                "lowest_expense_item":               lowest_expense_item,
                "lowest_expense_value":              lowest_expense_value,
                "most_frequent_purchased_inventory": most_frequent_purchased,
                "most_frequent_purchased_value":     most_frequent_value,
            },
            "reconciliation": {
                "opening_stock":    opening_stock,
                "purchases":        purchases_in_period,
                "closing_stock":    closing_stock,
                "cogs":             cogs,
                "of_which_wastage": wastage,
                "revenue":          revenue,
                "food_cost_pct":    food_cost_percentage,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch remaining inventory summary: {str(e)}",
        )


def _empty_summary_response(period, start_dt, end_dt):
    return {
        "success": True,
        "message": "Remaining inventory summary fetched successfully",
        "period": period.value,
        "period_range": {"start": start_dt.isoformat(), "end": end_dt.isoformat()},
        "has_activity": False,
        "data": {
            "opening_stock":                     0,
            "closing_stock":                     0,
            "purchases":                         0,
            "cogs":                              0,
            "wastage":                           0,
            "revenue":                           0,
            "food_cost_percentage":              0,
            "highest_expense_item":              None,
            "highest_expense_value":             None,
            "lowest_expense_item":               None,
            "lowest_expense_value":              None,
            "most_frequent_purchased_inventory": None,
            "most_frequent_purchased_value":     None,
        },
        "reconciliation": {
            "opening_stock":    0,
            "purchases":        0,
            "closing_stock":    0,
            "cogs":             0,
            "of_which_wastage": 0,
            "revenue":          0,
            "food_cost_pct":    0,
        },
    }


# ──────────────────────────────────────────────────────────────────────────────
# /itemwise
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/itemwise", status_code=status.HTTP_200_OK)
def get_itemwise_management(
    period: PeriodType = Query(PeriodType.MONTHLY),
    reference_date: Optional[date] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.tenant_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Tenant access required")

    if period != PeriodType.CUSTOM and (start_date or end_date):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="start_date and end_date are only valid when period=custom.",
        )

    start_dt, end_dt = get_period_range(period, reference_date, start_date, end_date)
    tenant_id = current_user.tenant_id

    try:
        items_query = (
            db.query(Inventory)
            .filter(Inventory.tenant_id == tenant_id, Inventory.is_active == True)
        )
        if search:
            items_query = items_query.filter(Inventory.name.ilike(f"%{search}%"))

        items    = items_query.all()
        item_ids = [i.id for i in items]

        if not item_ids:
            return {
                "success": True,
                "message": "Itemwise management data fetched successfully",
                "period": period.value,
                "period_range": {"start": start_dt.isoformat(), "end": end_dt.isoformat()},
                "total": 0,
                "data": [],
            }

        item_base_unit_map = {i.id: _resolve_unit(i.unit) for i in items}

        # ── Opening transactions (before period) ──────────────────────────────
        opening_txs = (
            db.query(InventoryTransaction)
            .filter(
                InventoryTransaction.tenant_id == tenant_id,
                InventoryTransaction.inventory_item_id.in_(item_ids),
                InventoryTransaction.transaction_date < start_dt,
            )
            .all()
        )

        # ── Period transactions ───────────────────────────────────────────────
        period_txs = (
            db.query(InventoryTransaction)
            .filter(
                InventoryTransaction.tenant_id == tenant_id,
                InventoryTransaction.inventory_item_id.in_(item_ids),
                InventoryTransaction.transaction_date >= start_dt,
                InventoryTransaction.transaction_date <= end_dt,
            )
            .all()
        )

        # ── Aggregate with unit normalisation ─────────────────────────────────
        opening_map = _aggregate_transactions_bulk(
            opening_txs, item_base_unit_map, positive_types=(TransactionType.PURCHASE,)
        )
        movement_map = _aggregate_transactions_bulk(
            period_txs, item_base_unit_map, positive_types=(TransactionType.PURCHASE,)
        )

        # ── Wastage map (quantity + value, always positive for display) ────────
        #  FIX: use _tx_value() for wastage value — consistent with summary
        wastage_map: dict[int, dict] = {}
        for tx in period_txs:
            if tx.transaction_type != TransactionType.WASTAGE:
                continue
            item_id   = tx.inventory_item_id
            base_unit = item_base_unit_map.get(item_id, "")
            tx_unit   = _resolve_unit(tx.unit) if hasattr(tx, "unit") and tx.unit else base_unit
            norm_qty  = _normalize_qty(float(tx.quantity or 0), tx_unit, base_unit)
            value     = _tx_value(tx)  #  was: float((tx.quantity or 0) * (tx.unit_cost or 0))
            if item_id not in wastage_map:
                wastage_map[item_id] = {"qty": 0.0, "value": 0.0}
            wastage_map[item_id]["qty"]   += norm_qty
            wastage_map[item_id]["value"] += value

        # ── Assemble result ───────────────────────────────────────────────────
        result = []
        for item in items:
            unit_label = _resolve_unit(item.unit)

            opening  = opening_map.get(item.id,  {"qty": 0.0, "value": 0.0})
            movement = movement_map.get(item.id, {"qty": 0.0, "value": 0.0})
            wastage  = wastage_map.get(item.id,  {"qty": 0.0, "value": 0.0})

            opening_qty   = max(opening["qty"],   0.0)
            opening_value = max(opening["value"], 0.0)

            wastage_qty   = wastage["qty"]
            wastage_value = wastage["value"]

            # Closing = opening + net movement (wastage already negative in movement)
            closing_qty   = max(opening_qty   + movement["qty"],   0.0)
            closing_value = max(opening_value + movement["value"], 0.0)

            result.append({
                "item_id":   item.id,
                "item_name": item.name,
                "unit":      unit_label,
                "opening": {
                    "quantity":         round(opening_qty, 4),
                    "quantity_display": f"{opening_qty:.2f} {unit_label}",
                    "value":            round(opening_value, 2),
                },
                "wastage": {
                    "quantity":         round(wastage_qty, 4),
                    "quantity_display": f"{wastage_qty:.2f} {unit_label}",
                    "value":            round(wastage_value, 2),
                },
                "closing": {
                    "quantity":         round(closing_qty, 4),
                    "quantity_display": f"{closing_qty:.2f} {unit_label}",
                    "value":            round(closing_value, 2),
                },
                "remaining_value": round(closing_value, 2),
            })

        return {
            "success": True,
            "message": "Itemwise management data fetched successfully",
            "period": period.value,
            "period_range": {
                "start": start_dt.isoformat(),
                "end":   end_dt.isoformat(),
            },
            "total": len(result),
            "data": result,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch itemwise management data: {str(e)}",
        )


# ──────────────────────────────────────────────────────────────────────────────
# /dashboard
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/dashboard", status_code=status.HTTP_200_OK)
def get_remaining_inventory_dashboard(
    period: PeriodType = Query(PeriodType.MONTHLY),
    reference_date: Optional[date] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(default=1, ge=1, description="Page number"),
    page_size: int = Query(default=10, ge=1, le=100, description="Rows per page"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Single endpoint that powers the entire Remaining Inventory dashboard.
    Combines /summary and /itemwise into one response.
    """
    summary = get_remaining_inventory_summary(
        period=period,
        reference_date=reference_date,
        start_date=start_date,
        end_date=end_date,
        db=db,
        current_user=current_user,
    )
    itemwise = get_itemwise_management(
        period=period,
        reference_date=reference_date,
        start_date=start_date,
        end_date=end_date,
        search=search,
        db=db,
        current_user=current_user,
    )

    # ── Pagination on itemwise data ───────────────────────────────────────────
    all_items = itemwise["data"]
    total     = len(all_items)
    start     = (page - 1) * page_size
    paged_items = all_items[start : start + page_size]

    return {
        "success": True,
        "message": "Remaining inventory dashboard fetched successfully",
        "period":       period.value,
        "period_range": summary["period_range"],
        "summary":      summary["data"],
        "itemwise": {
            "meta": {
                "total":       total,
                "page":        page,
                "page_size":   page_size,
                "total_pages": ceil(total / page_size) if total else 1,
                "search":      search,
            },
            "total": total,
            "data":  paged_items,
        },
    }


def _fmt_inr(value) -> str:
    return f"₹{float(value or 0):,.2f}"


def _period_label(start_dt: datetime, end_dt: datetime) -> str:
    s = start_dt.strftime("%d %b %Y")
    e = end_dt.strftime("%d %b %Y")
    return s if s == e else f"{s} – {e}"


def _fetch_report_payload(
    period, reference_date, start_date, end_date, search, db, current_user
) -> dict:
    """Call both internal functions and return a unified payload."""
    summary_resp = get_remaining_inventory_summary(
        period=period,
        reference_date=reference_date,
        start_date=start_date,
        end_date=end_date,
        db=db,
        current_user=current_user,
    )
    itemwise_resp = get_itemwise_management(
        period=period,
        reference_date=reference_date,
        start_date=start_date,
        end_date=end_date,
        search=search,
        db=db,
        current_user=current_user,
    )
    return {
        "summary":      summary_resp["data"],
        "period_range": summary_resp["period_range"],
        "items":        itemwise_resp["data"],
    }

@router.get("/download/excel", status_code=200)
def download_remaining_inventory_excel(
    period: PeriodType = Query(PeriodType.DAILY),
    reference_date: Optional[date] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.tenant_id:
        raise HTTPException(status_code=403, detail="Tenant access required")

    payload = _fetch_report_payload(
        period, reference_date, start_date, end_date, search, db, current_user
    )
    summary      = payload["summary"]
    items        = payload["items"]
    period_range = payload["period_range"]

    start_dt = datetime.fromisoformat(period_range["start"])
    end_dt   = datetime.fromisoformat(period_range["end"])
    label    = _period_label(start_dt, end_dt)

    # ── Style constants ───────────────────────────────────────────────────────
    ORANGE      = "FF6B35"
    LIGHT_ORANGE= "FFF3EE"
    GREY        = "F5F5F5"
    MID_GREY    = "E0E0E0"
    WHITE       = "FFFFFF"

    hdr_font   = Font(name="Calibri", bold=True, color=WHITE, size=10)
    bold_font  = Font(name="Calibri", bold=True, size=10)
    norm_font  = Font(name="Calibri", size=10)
    small_font = Font(name="Calibri", size=9, color="555555")

    orange_fill      = PatternFill("solid", fgColor=ORANGE)
    light_fill       = PatternFill("solid", fgColor=LIGHT_ORANGE)
    grey_fill        = PatternFill("solid", fgColor=GREY)
    mid_grey_fill    = PatternFill("solid", fgColor=MID_GREY)
    white_fill       = PatternFill("solid", fgColor=WHITE)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def _c(ws, row, col, value="", font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:  cell.font   = font
        if fill:  cell.fill   = fill
        if align: cell.alignment = align
        cell.border = thin_border()
        return cell

    wb = Workbook()
    ws = wb.active
    ws.title = "Remaining Inventory"

    # ── Title block ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value     = "Remaining Inventory Report"
    t.font      = Font(name="Calibri", bold=True, size=16, color=ORANGE)
    t.alignment = center
    t.fill      = light_fill
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    d = ws["A2"]
    d.value = f"Period: {label}   |   Generated: {datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y, %I:%M %p')}"
    d.font      = Font(name="Calibri", size=9, color="777777")
    d.alignment = center
    ws.row_dimensions[2].height = 16

    # ── Summary header ────────────────────────────────────────────────────────
    ws.merge_cells("A4:J4")
    sh = ws["A4"]
    sh.value     = "Summary"
    sh.font      = Font(name="Calibri", bold=True, size=11, color=ORANGE)
    sh.fill      = light_fill
    sh.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 20

    summary_cols = [
        ("Opening Stock",    _fmt_inr(summary.get("opening_stock"))),
        ("Closing Stock",    _fmt_inr(summary.get("closing_stock"))),
        ("Purchases",        _fmt_inr(summary.get("purchases"))),
        ("COGS",             _fmt_inr(summary.get("cogs"))),
        ("Wastage",          _fmt_inr(summary.get("wastage"))),
        ("Revenue",          _fmt_inr(summary.get("revenue"))),
        ("Food Cost %",      f"{summary.get('food_cost_percentage', 0):.2f}%"),
        ("Highest Expense",  summary.get("highest_expense_item") or "-"),
        ("Lowest Expense",   summary.get("lowest_expense_item")  or "-"),
        ("Most Frequent",    summary.get("most_frequent_purchased_inventory") or "-"),
    ]

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 24
    for col_idx, (header, value) in enumerate(summary_cols, start=1):
        _c(ws, 5, col_idx, header, font=small_font,  fill=mid_grey_fill, align=center)
        _c(ws, 6, col_idx, value,  font=bold_font,   fill=light_fill,    align=center)

    # ── Table header ──────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 24
    table_headers = [
        "Item", "Opening Qty", "Unit",
        "Wastage Qty", "Wastage Value",
        "Closing Qty", "Closing Value",
        "Opening Value", "Remaining Value", "Wastage %",
    ]
    for col_idx, h in enumerate(table_headers, start=1):
        _c(ws, 8, col_idx, h, font=hdr_font, fill=orange_fill, align=center)

    # ── Table rows ────────────────────────────────────────────────────────────
    for row_offset, item in enumerate(items):
        row     = 9 + row_offset
        row_fill = white_fill if row_offset % 2 == 0 else grey_fill
        ws.row_dimensions[row].height = 18

        opening_val  = item["opening"]["value"]
        closing_val  = item["closing"]["value"]
        wastage_val  = item["wastage"]["value"]
        wastage_qty  = item["wastage"]["quantity"]
        opening_qty  = item["opening"]["quantity"]
        closing_qty  = item["closing"]["quantity"]
        unit         = item["unit"]
        wastage_pct  = round((wastage_val / opening_val) * 100, 1) if opening_val > 0 else 0.0

        row_data = [
            (item["item_name"],              left,   norm_font),
            (f"{opening_qty:.2f}",           center, norm_font),
            (unit,                           center, norm_font),
            (f"{wastage_qty:.2f}",           center, norm_font),
            (_fmt_inr(wastage_val),          right,  norm_font),
            (f"{closing_qty:.2f}",           center, norm_font),
            (_fmt_inr(closing_val),          right,  norm_font),
            (_fmt_inr(opening_val),          right,  norm_font),
            (_fmt_inr(item["remaining_value"]), right, bold_font),
            (f"{wastage_pct:.1f}%",          center, norm_font),
        ]

        for col_idx, (value, align, font) in enumerate(row_data, start=1):
            _c(ws, row, col_idx, value, font=font, fill=row_fill, align=align)

    # ── Totals row ────────────────────────────────────────────────────────────
    total_row = 9 + len(items)
    ws.row_dimensions[total_row].height = 22
    total_opening  = sum(i["opening"]["value"]   for i in items)
    total_wastage  = sum(i["wastage"]["value"]    for i in items)
    total_closing  = sum(i["closing"]["value"]    for i in items)
    total_remaining= sum(i["remaining_value"]     for i in items)

    totals = [
        ("TOTAL",                  left,   bold_font),
        ("",                       center, bold_font),
        ("",                       center, bold_font),
        ("",                       center, bold_font),
        (_fmt_inr(total_wastage),  right,  bold_font),
        ("",                       center, bold_font),
        (_fmt_inr(total_closing),  right,  bold_font),
        (_fmt_inr(total_opening),  right,  bold_font),
        (_fmt_inr(total_remaining),right,  bold_font),
        ("",                       center, bold_font),
    ]
    for col_idx, (value, align, font) in enumerate(totals, start=1):
        _c(ws, total_row, col_idx, value, font=font, fill=mid_grey_fill, align=align)

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, w in enumerate([28, 13, 10, 13, 16, 13, 16, 16, 17, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Stream ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"remaining_inventory_{start_dt.date()}_{end_dt.date()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
